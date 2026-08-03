"""Parse a bulk team-name upload (CSV or .xlsx) for the Add Team section.

CSV uses the stdlib `csv` module; .xlsx is read with openpyxl (already a
project dependency — see tournaments/roster_import.py, which reads the same
file types for the basketball roster import feature). Kept as its own module
rather than sharing code with roster_import.py so the two upload features
stay fully independent — this one is scoped to team names only.

`parse_teams(uploaded_file)` returns a `TeamImportParse` with:
  - names: list of valid team-name strings, ready for Team.objects.create()
  - errors: human-readable strings for an invalid file / invalid rows
  - skipped_blank: count of empty rows ignored gracefully
"""
import csv
import io
from dataclasses import dataclass, field

import openpyxl

# Accepted header labels (matched case-insensitively, trimmed).
_NAME_HEADERS = {'team name', 'team', 'name'}


@dataclass
class TeamImportParse:
    names: list = field(default_factory=list)
    errors: list = field(default_factory=list)
    skipped_blank: int = 0


def parse_teams(uploaded_file):
    """Dispatch on file extension, then validate the tabular rows."""
    name = (getattr(uploaded_file, 'name', '') or '').lower()
    try:
        if name.endswith('.csv'):
            rows = _read_csv_rows(uploaded_file)
        elif name.endswith('.xlsx'):
            rows = _read_xlsx_rows(uploaded_file)
        else:
            return TeamImportParse(errors=['Unsupported file type — upload a .csv or .xlsx file.'])
    except Exception:
        # Corrupt file, bad encoding, unreadable sheet — never fail silently.
        return TeamImportParse(errors=['Could not read the file. Make sure it is a valid CSV or .xlsx file.'])

    return _validate_rows(rows)


# ----------------------------------------------------------------------
# Readers — each returns a list of rows, every row a list of cell strings
# ----------------------------------------------------------------------
def _read_csv_rows(uploaded_file):
    raw = uploaded_file.read()
    if isinstance(raw, bytes):
        # utf-8-sig strips a BOM that Excel commonly writes on CSV export.
        text = raw.decode('utf-8-sig', errors='replace')
    else:
        text = raw
    return [list(r) for r in csv.reader(io.StringIO(text))]


def _cell_str(value):
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _read_xlsx_rows(uploaded_file):
    wb = openpyxl.load_workbook(io.BytesIO(uploaded_file.read()), read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        return [[_cell_str(c) for c in row] for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


# ----------------------------------------------------------------------
# Validation
# ----------------------------------------------------------------------
def _validate_rows(rows):
    result = TeamImportParse()

    # First non-empty row is the header.
    header_idx = None
    for i, row in enumerate(rows):
        if any((c or '').strip() for c in row):
            header_idx = i
            break
    if header_idx is None:
        result.errors.append('The file is empty.')
        return result

    header = [(c or '').strip().lower() for c in rows[header_idx]]
    name_col = next((i for i, h in enumerate(header) if h in _NAME_HEADERS), None)
    if name_col is None:
        result.errors.append('Missing required column: Team Name. '
                             'Expected a header named "Team Name".')
        return result

    seen = set()
    for i in range(header_idx + 1, len(rows)):
        row = rows[i]
        excel_row = i + 1  # 1-based, for messages

        if not any((c or '').strip() for c in row):
            result.skipped_blank += 1
            continue

        team_name = (row[name_col].strip() if name_col < len(row) else '')
        if not team_name:
            result.errors.append(f'Row {excel_row}: team name is empty.')
            continue

        key = team_name.casefold()
        if key in seen:
            result.errors.append(f'Row {excel_row}: duplicate team "{team_name}" in the file — skipped.')
            continue
        seen.add(key)

        result.names.append(team_name)

    return result
