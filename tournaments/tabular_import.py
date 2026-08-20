"""Shared low-level tabular readers for bulk-import features that accept
either a file upload (.csv/.xlsx) or pasted text.

team_import.py and roster_import.py deliberately keep their own copies of the
file-reading logic (see their docstrings) so this module is not used by them.
It exists for participant_import.py, which additionally needs a pasted-text
reader that neither of those older modules has.
"""
import csv
import io

import openpyxl


def _cell_str(value):
    """openpyxl cell value -> trimmed string. Integers come back as ints
    (jersey 10, not 10.0), so render whole floats without a trailing '.0'."""
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _read_csv_rows(uploaded_file):
    raw = uploaded_file.read()
    if isinstance(raw, bytes):
        # utf-8-sig strips a BOM that Excel commonly writes on CSV export.
        text = raw.decode('utf-8-sig', errors='replace')
    else:
        text = raw
    return [list(r) for r in csv.reader(io.StringIO(text))]


def _read_xlsx_rows(uploaded_file):
    # read_only + data_only: stream cells and take computed values, not formulas.
    wb = openpyxl.load_workbook(io.BytesIO(uploaded_file.read()),
                                read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        return [[_cell_str(c) for c in row] for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


def read_uploaded_file_rows(uploaded_file):
    """Dispatch on file extension. Returns a list of rows (each a list of
    cell strings), or raises ValueError('unsupported_type' | 'unreadable')."""
    name = (getattr(uploaded_file, 'name', '') or '').lower()
    try:
        if name.endswith('.csv'):
            return _read_csv_rows(uploaded_file)
        elif name.endswith('.xlsx'):
            return _read_xlsx_rows(uploaded_file)
        raise ValueError('unsupported_type')
    except ValueError:
        raise
    except Exception:
        # Corrupt file, bad encoding, unreadable sheet — never fail silently.
        raise ValueError('unreadable')


def read_pasted_text_rows(text):
    """Split pasted text into rows. Sniffs once for the whole paste whether
    columns are tab- or comma-separated (tab wins if any line has one), so a
    stray comma inside a name/phone cell doesn't misparse a tab-delimited
    paste."""
    lines = text.replace('\r\n', '\n').replace('\r', '\n').split('\n')
    # A trailing newline is a paste artifact, not a user-intended blank row.
    if lines and lines[-1] == '':
        lines.pop()
    delimiter = '\t' if any('\t' in line for line in lines) else ','
    reader = csv.reader(lines, delimiter=delimiter)
    return [list(r) for r in reader]
