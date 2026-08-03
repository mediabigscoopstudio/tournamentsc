"""Parse a basketball roster upload (CSV or .xlsx) into validated player rows.

CSV uses the stdlib `csv` module; .xlsx is read with openpyxl.

`parse_roster(uploaded_file)` returns a `RosterParse` with:
  - members: list of {'name': str, 'jersey': str} ready for TeamMembership
  - errors:  human-readable strings for invalid file / invalid rows
  - skipped_blank: count of empty rows ignored gracefully
"""
import csv
import io
from dataclasses import dataclass, field

import openpyxl

# Accepted header labels (matched case-insensitively, trimmed).
_NAME_HEADERS = {'player name', 'name', 'player', 'full name'}
_JERSEY_HEADERS = {'jersey number', 'jersey', 'jersey no', 'jersey #', 'number', 'no', '#'}

MAX_JERSEY = 99


@dataclass
class RosterParse:
    members: list = field(default_factory=list)
    errors: list = field(default_factory=list)
    skipped_blank: int = 0


def parse_roster(uploaded_file):
    """Dispatch on file extension, then validate the tabular rows."""
    name = (getattr(uploaded_file, 'name', '') or '').lower()
    try:
        if name.endswith('.csv'):
            rows = _read_csv_rows(uploaded_file)
        elif name.endswith('.xlsx'):
            rows = _read_xlsx_rows(uploaded_file)
        else:
            return RosterParse(errors=['Unsupported file type — upload a .csv or .xlsx file.'])
    except Exception:
        # Corrupt CSV/zip, bad encoding, unreadable sheet — never fail silently.
        return RosterParse(errors=['Could not read the file. Make sure it is a valid CSV or .xlsx file.'])

    return _validate_rows(rows)


# ----------------------------------------------------------------------
# Readers — each returns a list of rows, every row a list of cell strings
# ----------------------------------------------------------------------
def _read_csv_rows(uploaded_file):
    raw = uploaded_file.read()
    if isinstance(raw, bytes):
        # utf-8-sig strips a BOM that Excel commonly writes on CSV export.
        text = raw.decode('utf-8-sig', errors='replace')
    else:
        text = raw
    return [list(r) for r in csv.reader(io.StringIO(text))]


def _cell_str(value):
    """openpyxl cell value -> trimmed string. Integers come back as ints
    (jersey 10, not 10.0), so render whole floats without a trailing '.0'."""
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _read_xlsx_rows(uploaded_file):
    # read_only + data_only: stream cells and take computed values, not formulas.
    wb = openpyxl.load_workbook(io.BytesIO(uploaded_file.read()),
                                read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        return [[_cell_str(c) for c in row] for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


# ----------------------------------------------------------------------
# Validation
# ----------------------------------------------------------------------
def _norm_jersey(raw):
    """Return (value, error). Empty is allowed (matches manual add)."""
    s = str(raw or '').strip()
    if not s:
        return '', None
    try:
        f = float(s)          # tolerate '10' and Excel's '10.0'
    except ValueError:
        return None, f'jersey number "{s}" is not a number'
    if f != int(f):
        return None, f'jersey number "{s}" is not a whole number'
    n = int(f)
    if n < 0 or n > MAX_JERSEY:
        return None, f'jersey number {n} is out of range (0–{MAX_JERSEY})'
    return str(n), None


def _validate_rows(rows):
    result = RosterParse()

    # First non-empty row is the header.
    header_idx = None
    for i, row in enumerate(rows):
        if any((c or '').strip() for c in row):
            header_idx = i
            break
    if header_idx is None:
        result.errors.append('The file is empty.')
        return result

    header = [(c or '').strip().lower() for c in rows[header_idx]]
    name_col = next((i for i, h in enumerate(header) if h in _NAME_HEADERS), None)
    jersey_col = next((i for i, h in enumerate(header) if h in _JERSEY_HEADERS), None)

    missing = []
    if name_col is None:
        missing.append('Player Name')
    if jersey_col is None:
        missing.append('Jersey Number')
    if missing:
        result.errors.append('Missing required column(s): ' + ', '.join(missing) +
                             '. Expected headers "Player Name" and "Jersey Number".')
        return result

    seen = set()
    for i in range(header_idx + 1, len(rows)):
        row = rows[i]
        excel_row = i + 1  # 1-based, for messages

        if not any((c or '').strip() for c in row):
            result.skipped_blank += 1
            continue

        name = (row[name_col].strip() if name_col < len(row) else '')
        jersey_raw = (row[jersey_col] if jersey_col < len(row) else '')

        if not name:
            result.errors.append(f'Row {excel_row}: player name is empty.')
            continue

        jersey, jerr = _norm_jersey(jersey_raw)
        if jerr:
            result.errors.append(f'Row {excel_row}: {jerr}.')
            continue

        key = name.casefold()
        if key in seen:
            result.errors.append(f'Row {excel_row}: duplicate player "{name}" in the file — skipped.')
            continue
        seen.add(key)

        result.members.append({'name': name, 'jersey': jersey})

    return result
